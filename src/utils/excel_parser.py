#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel/CSV data parser.
"""

from pathlib import Path
from typing import List, Dict

from openpyxl import load_workbook

from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelParser:
    """Parse Excel/CSV data files."""

    EXPECTED_COLUMNS = {
        'product_id',
        'product_name',
        'category',
        'selling_points',
        'price',
        'image_path',
        'style',
    }

    def parse(self, file_path: Path) -> List[Dict]:
        """Parse Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of product dictionaries
        """
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            return self._parse_excel(file_path)
        elif file_path.suffix.lower() == '.csv':
            return self._parse_csv(file_path)
        else:
            logger.error(f'Unsupported file format: {file_path.suffix}')
            return []

    def _parse_excel(self, file_path: Path) -> List[Dict]:
        """Parse Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of product dictionaries
        """
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Get headers
            headers = [cell.value for cell in sheet[1]]
            products = []

            # Parse rows
            for row in sheet.iter_rows(min_row=2, values_only=False):
                product_data = {}
                for idx, cell in enumerate(row):
                    header = headers[idx]
                    value = cell.value

                    # Parse comma-separated selling points
                    if header == 'selling_points' and value:
                        value = [s.strip() for s in str(value).split(',')]

                    product_data[header] = value

                # Skip empty rows
                if product_data.get('product_id'):
                    products.append(product_data)

            logger.info(f'Parsed {len(products)} products from Excel')
            return products

        except Exception as e:
            logger.error(f'Failed to parse Excel file: {e}')
            return []

    def _parse_csv(self, file_path: Path) -> List[Dict]:
        """Parse CSV file.

        Args:
            file_path: Path to CSV file

        Returns:
            List of product dictionaries
        """
        try:
            import csv
            products = []

            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Parse comma-separated selling points
                    if 'selling_points' in row and row['selling_points']:
                        row['selling_points'] = [s.strip() for s in row['selling_points'].split(',')]

                    products.append(row)

            logger.info(f'Parsed {len(products)} products from CSV')
            return products

        except Exception as e:
            logger.error(f'Failed to parse CSV file: {e}')
            return []
